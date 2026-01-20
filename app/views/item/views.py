import json

from django.db import connection
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

from app.models.item import Item
from app.views.item.form import ItemForm
import requests
import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


def home(request):
    items = Item.objects.all()
    form = ItemForm()
    context = {
        'items': items,
        'form': form,
    }

    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            new_item = form.save()
            return JsonResponse({
                'item_id': new_item.item_id,
                'item_no': new_item.item_no,
                'item_name': new_item.item_name,
                'stock': new_item.stock,
                'harga': new_item.harga,
            })
        else:
            return JsonResponse({'error': 'Data tidak valid'}, status=400)
    else:
        return render(request, "item/home.html", context)

def home_vue(request):
    items = Item.objects.all()
    form = ItemForm()
    context = {
        'items': items,
        'form': form,
    }

    if request.method == "POST":
        form = ItemForm(request.POST)
        if form.is_valid():
            new_item = form.save()
            return JsonResponse({
                'item_id': new_item.item_id,
                'item_no': new_item.item_no,
                'item_name': new_item.item_name,
                'stock': new_item.stock,
                'harga': new_item.harga,
            })
        else:
            return JsonResponse({'error': 'Data tidak valid'}, status=400)
    else:
        return render(request, "item/home_vue.html", context)
def fetch_items(request):
    items = Item.objects.all()  # Ambil semua item dari database
    items_data = list(items.values('item_id', 'item_no', 'item_name', 'stock', 'harga'))
    return JsonResponse(items_data, safe=False)

def detail(request, item_id):
    item = Item.objects.get(item_id=item_id)
    form = ItemForm(instance=item)
    if request.method == "POST":
        form = ItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return redirect("item_home")

    context = {
        'item': item,
        'form': form
    }
    return render(request, "item/form.html", context)

def delete(request, item_id):
    with connection.cursor() as cursor:
        cursor.execute("CALL log_delete(%s)", [item_id])
    return JsonResponse({'status': 'success', 'item_id': item_id})

def requestView(request):
    url = "http://www.omdbapi.com/?apikey=93a071b5&s=batman&page=1"
    response = requests.get(url, timeout=5)
    data = response.json()

    context = {
        'items': data.get('Search', [])
    }

    return render(request, "item/request_view.html", context)

def export_items(request):
    data = json.loads(request.body)
    items = data.get('items', [])

    df = pd.DataFrame(items)
    df = df.drop(columns=['item_id', 'created_at', 'updated_at'], errors='ignore')

    df['total'] = df['stock'].astype(int) * df['harga'].astype(int)

    df = df.rename(columns={
        'item_no': 'Nomor Produk',
        "item_name" 'Nama Produk'
        'stock': 'Stok',
        'harga': 'Harga',
        'total': 'Total'
    })

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=laporan_items.xlsx'

    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=3, sheet_name='Laporan')

        ws = writer.sheets['Laporan']

        red_fill = PatternFill(
            start_color='FF0000',
            end_color='FF0000',
            fill_type='solid'
        )

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
        ws['A1'] = 'LAPORAN ITEM'
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
        ws['A2'] = 'Export dari Vue + Django (Pandas + Openpyxl)'
        ws['A2'].alignment = Alignment(horizontal='center')

        for cell in ws[4]:
            cell.font = Font(bold=True)

        for row in ws.iter_rows(min_row=5, min_col=3, max_col=5):
            for cell in row:
                cell.number_format = '#,##0'

        for row in ws.iter_rows(min_row=5, min_col=3, max_col=3):
            cell = row[0]
            is_empty = cell.value in (None, "", " ")
            is_low_stock = isinstance(cell.value, (int, float)) and cell.value < 5

            if is_empty or is_low_stock:
                cell.fill = red_fill

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

    return response